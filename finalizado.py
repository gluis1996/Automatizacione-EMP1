#!/usr/bin/env python
# coding: utf-8

# In[25]:


import pandas as pd
from openpyxl import load_workbook
import numpy as np
import glob
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from pathlib import Path
pd.set_option('display.expand_frame_repr', False)  # Evita que las filas se dividan
pd.set_option('display.max_columns', None)  # Muestra todas las columnas


# ## Obtencion de las rutas de los archivos implicados

# In[26]:


wb = load_workbook('config.xlsx')
ws = wb.active
rutaExcelCRM = ws.cell(row=3, column=2).value
excelListaPrecios =  ws.cell(row=4, column=2).value
excelListaPreciosLG =  ws.cell(row=5, column=2).value
print(rutaExcelCRM)
print(excelListaPrecios)
print(excelListaPreciosLG)


# ## Funciones

# In[27]:


# Funcion que Procesara el excel de Lista precio 
def procesarExcelListaPrecio(ruta_excel, nombre_hoja):
    # 1. Leer hoja sin cabecera
    df = pd.read_excel(ruta_excel, header=None, sheet_name=nombre_hoja)
    df_str = df.astype(str).apply(lambda col: col.str.upper().str.strip())

    # 2. Buscar PRIMERA fila que parezca cabecera
    claves_obligatorias = ["CÓDIGO", "CODIGO", "NOMBRE COMERCIAL"]
    fila_cabecera = None

    for i, fila in df_str.iterrows():
        textos = list(fila.values)
        tiene_codigo = any("CÓDIGO" in t or "CODIGO" in t for t in textos)
        tiene_nombre = any("NOMBRE COMERCIAL" in t for t in textos)
        if tiene_codigo and tiene_nombre:
            fila_cabecera = i
            break

    if fila_cabecera is None:
        raise Exception("❌ No se encontró cabecera con Código y Nombre comercial.")

    # Cabecera oficial (la primera que encontramos)
    columnas_raw = df.iloc[fila_cabecera].tolist()

    registros = []

    # 3. Recorrer filas de datos desde la siguiente fila
    for i in range(fila_cabecera + 1, len(df)):
        fila = df.iloc[i]
        fila_str = df_str.iloc[i]
        celdas_no_nulas = fila.dropna()

        # --- saltar filas totalmente vacías ---
        if celdas_no_nulas.empty:
            continue

        # --- saltar filas que son solo un número (totales) ---
        if len(celdas_no_nulas) == 1 and isinstance(celdas_no_nulas.iloc[0], (int, float)):
            continue

        textos = [t for t in fila_str.values if t != ""]
        texto_unico = textos[0] if len(textos) == 1 else ""

        # --- saltar títulos / categorías / totales ---
        if len(textos) == 1:
            t = texto_unico
            if any(p in t for p in ["LISTA DE PRECIO", "LAVADORA", "SECADORA", "PRENSA", "CABINA", "TOTAL"]):
                continue

        # --- saltar cabeceras repetidas (Ítem, Código, Nombre comercial, etc.) ---
        if any(t.startswith("ÍTEM") or t == "ITEM" for t in textos) and \
           any("CÓDIGO" in t or "CODIGO" in t for t in textos):
            continue

        # 4. Convertir a diccionario usando SIEMPRE la primera cabecera
        dic = dict(zip(columnas_raw, fila.tolist()))
        registros.append(dic)

    # 5. DataFrame final con TODAS las columnas
    df_full = pd.DataFrame(registros)

    # 6. Normalizar nombres de columnas
    df_full.columns = df_full.columns.astype(str).str.upper().str.strip()

    # 7. Mapear a las columnas que te interesan
    #    (por si cambian ligeramente los nombres entre hojas)
    mapa_columnas = {}

    # Código
    for col in ["CÓDIGO", "CODIGO"]:
        if col in df_full.columns:
            mapa_columnas["CÓDIGO"] = col
            break

    # Nombre comercial
    for col in ["NOMBRE COMERCIAL"]:
        if col in df_full.columns:
            mapa_columnas["NOMBRE COMERCIAL"] = col
            break

    # Costo
    for col in ["COSTO COMPASS", "COSTO COMPRA", "COSTO OCTUBRE 2025"]:
        if col in df_full.columns:
            mapa_columnas["COSTO COMPASS"] = col
            break

    # Nos quedamos solo con esas columnas
    df_final = df_full[list(mapa_columnas.values())]
    df_final.columns = list(mapa_columnas.keys())

    # Limpiar filas con código vacío
    df_final = df_final[df_final["CÓDIGO"].notna()]
    df_final = df_final.dropna(how='all', subset=["CÓDIGO", "NOMBRE COMERCIAL", "COSTO COMPASS"])
    return df_final


# In[28]:


# Funcion que Procesara el excel de lista precios LG
def procesarExcelListaPrecio_LG(ruta_excel, nombre_hoja):
    # 1) Leer la hoja COMPLETA sin cabecera
    df = pd.read_excel(ruta_excel, header=None, sheet_name=nombre_hoja)

    # Versión en texto (mayúsculas) para poder buscar palabras clave
    df_str = df.astype(str).apply(lambda col: col.str.upper().str.strip())

    # =====================================================
    # 2) Detectar la PRIMERA fila que será la cabecera base
    #    (la que tenga CÓDIGO y DESCRIPCION / NOMBRE COMERCIAL)
    # =====================================================
    fila_cabecera = None

    for i, fila in df_str.iterrows():
        textos = list(fila.values)

        # ¿Hay alguna celda que contenga "CÓDIGO" o "CODIGO"?
        tiene_codigo = any("CÓDIGO" in t or "CODIGO" in t for t in textos)

        # ¿Hay alguna celda que contenga "DESCRIPCION" o "NOMBRE COMERCIAL"?
        tiene_nombre = any(
            "DESCRIPCION" in t or "DESCRIPCIÓN" in t or "NOMBRE COMERCIAL" in t
            for t in textos
        )

        # Primera fila que cumple ambas condiciones → cabecera oficial
        if tiene_codigo and tiene_nombre:
            fila_cabecera = i
            break

    if fila_cabecera is None:
        raise Exception("❌ No se encontró cabecera válida.")

    # Guardamos los nombres originales de la cabecera
    columnas_raw = df.iloc[fila_cabecera].tolist()

    registros = []

    # =====================================================
    # 3) Recorrer TODAS las filas de datos que vienen después
    #    y mapearlas SIEMPRE contra la PRIMERA cabecera
    # =====================================================
    for i in range(fila_cabecera + 1, len(df)):
        fila = df.iloc[i]        # fila con tipos reales (número, fecha, etc.)
        fila_str = df_str.iloc[i]  # la misma fila pero en texto upper()

        # Si la fila está completamente vacía → se ignora
        if fila.dropna().empty:
            continue

        textos = [t for t in fila_str.values if t != ""]

        # Filtrar títulos como "SECADORA ...", "LAVADORA ...", "TOTAL", etc.
        if len(textos) == 1:
            if any(p in textos[0] for p in
                   ["SECADORA", "LAVADORA", "TITAN", "MAX", "TOTAL", "STOCK", "GIANT"]):
                continue

        # Si pasó los filtros, construimos un dict usando SIEMPRE la primera cabecera
        registros.append(dict(zip(columnas_raw, fila.tolist())))

    # Pasamos la lista de diccionarios a DataFrame
    df_full = pd.DataFrame(registros)

    # Normalizamos nombres de columnas: texto, mayúsculas, sin espacios
    df_full.columns = df_full.columns.astype(str).str.upper().str.strip()

    # =====================================================
    # 4) Mapeo flexible de columnas
    #    (por si cambian ligeramente entre hojas)
    # =====================================================
    mapa = {}  # aquí guardamos: nombre_estándar -> nombre_real_en_df_full

    # Columna "CÓDIGO"
    for col in ["CÓDIGO", "CODIGO"]:
        if col in df_full.columns:
            mapa["CÓDIGO"] = col
            break

    # Columna "DESCRIPCION" (acepta varias variantes)
    for col in ["DESCRIPCION", "DESCRIPCIÓN", "NOMBRE COMERCIAL"]:
        if col in df_full.columns:
            mapa["DESCRIPCION"] = col
            break

    # Columna de costo (prioridad: COSTO COMPASS, luego COSTO COMPRA, luego COSTO ACTUAL)
    for col in ["COSTO COMPASS", "COSTO COMPRA", "COSTO ACTUAL"]:
        if col in df_full.columns:
            mapa["COSTO COMPASS"] = col
            break

    # Nos quedamos SOLO con las columnas mapeadas
    df_final = df_full[list(mapa.values())]

    # Renombramos las columnas al nombre estándar: CÓDIGO, DESCRIPCION, COSTO COMPASS
    df_final.columns = list(mapa.keys())

    # =====================================================
    # 5) LIMPIEZA FINAL DE FILAS
    #    - Quitar filas vacías
    #    - Quitar títulos/categorías sin código
    #    - Quitar filas donde DESCRIPCION y COSTO estén vacíos
    # =====================================================

    # Quitar filas donde TODAS las columnas estén NaN
    df_final = df_final.dropna(how="all")

    # Aseguramos que CÓDIGO y DESCRIPCION sean string para poder evaluar
    df_final["CÓDIGO"] = df_final["CÓDIGO"].astype(str).str.strip()
    df_final["DESCRIPCION"] = df_final["DESCRIPCION"].astype(str).str.strip()

    # 5.1 Quitar filas donde el CÓDIGO no es válido (vacío, NaN, etc.)
    mask_codigo_valido = df_final["CÓDIGO"].str.match(r"^[A-Za-z0-9]+$", na=False)
    df_final = df_final[mask_codigo_valido]

    # 5.2 Quitar filas donde DESCRIPCION Y COSTO están vacíos a la vez
    desc_vacia = df_final["DESCRIPCION"].isna() | (df_final["DESCRIPCION"] == "")
    costo_vacio = df_final["COSTO COMPASS"].isna()
    mask_mantener = ~(desc_vacia & costo_vacio)   # mantener todo lo que NO tenga ambas vacías

    df_final = df_final[mask_mantener]

    # 5.3 Resetear índice para que quede limpio
    df_final = df_final.reset_index(drop=True)

    return df_final


# In[29]:


# Funcion que Procesara los datos del excel de CRM en la hoja CRM
def procesarCRM(path, hoja_principal="Base CRM", hoja_backup="Hoja1"):
    """
    Lee un Excel detectando:
    - qué hoja usar (principal o backup)
    - primera fila válida como encabezado
    - limpia filas vacías
    - estandariza nombres de columnas
    """

    column_map = {
        "No.": "numero",
        "No": "numero",
        "N°": "numero",
        "nro": "numero",

        "Fecha": "fecha",
        "Asesor": "asesor",
        "Contacto": "contacto",
        "Cargo": "cargo",

        "teléfono": "telefono",
        "telefono": "telefono",
        "Teléfono": "telefono",

        "Mail": "mail",
        "Correo": "mail",

        "Razón Social": "razon_social",
        "Razon Social": "razon_social",

        "RUC": "ruc",
        "Web": "web",
        "Departamento": "departamento",
        "Distrito": "distrito",
        "Rubro": "rubro",
        "Tipo de producto": "tipo_producto",
        "Tipo de Producto": "tipo_producto",

        "Origen del lead": "origen_lead",
        "Descripción": "descripcion",
        "Estado": "estado",
        "Monto de la oportunidad": "monto_oportunidad",
        "No. Cotización": "numero_cotizacion",

        "Valor Cotización": "valor_cotizacion",
        "Valor Cotización + IGV": "valor_cotizacion",

        "Fecha seguimiento": "fecha_seguimiento",
        "Semana": "semana",
    }

    try:
        with pd.ExcelFile(path) as xls:
            if hoja_principal in xls.sheet_names:
                hoja = hoja_principal
            elif hoja_backup in xls.sheet_names:
                hoja = hoja_backup
            else:
                print(f"❌ Ninguna hoja válida encontrada en: {path}")
                return None

            df_raw = pd.read_excel(xls, sheet_name=hoja, header=None)
    except Exception as e:
        print(f"❌ Error al abrir archivo {path}: {e}")
        return None

    header_row = next((i for i in range(len(df_raw)) if df_raw.iloc[i].notna().any()), None) # busca la primera fila que no esta vacia

    if header_row is None:
        print(f"❌ No se encontró encabezado en: {path}")
        return None

    df = pd.read_excel(path, sheet_name=hoja, header=header_row)

    df = df.dropna(how='all').reset_index(drop=True) # quita las fila completamente vacias

    # Estandarizar columnas
    df.columns = df.columns.astype(str).str.strip()  # quita espacios

    df = df.rename(columns=lambda c: column_map.get(c, c.lower().replace(" ", "_")))

    return df


# ## Creacion de consolidado de Lista precios

# In[30]:


# Usamos la funcion para crear el dataframe de lista productos con el excel lista productos
df_a = pd.ExcelFile(excelListaPrecios)
hojas = df_a.sheet_names
concat_a = []
for h in hojas:
    df = procesarExcelListaPrecio(excelListaPrecios,h)
    concat_a.append(df)
sd = pd.concat(concat_a, ignore_index=True)
sd = sd.dropna(how='all')
sd = sd.replace(r'^\s*$', np.nan, regex=True)
sd = sd.dropna(how='all', subset=["CÓDIGO", "NOMBRE COMERCIAL", "COSTO COMPASS"])
sd['COSTO COMPASS'] = sd['COSTO COMPASS'].fillna(0)
resExcel1 = sd.copy()


# In[31]:


# Usamos la funcion para crear el dataframe de lista productos con el excel lista productos Lg
resExcel2 = procesarExcelListaPrecio_LG(excelListaPreciosLG, 'RESUMEN GENERAL LG')
resExcel2 = resExcel2.rename(columns={'CÓDIGO':'CÓDIGO','DESCRIPCION':'NOMBRE COMERCIAL'})


# In[32]:


# Unimos ambos dataframe en uno solo
dfExel1 = resExcel1.copy()
dfExcel2 = resExcel2.copy()
dfFinalListaPrecios = pd.concat([dfExel1,dfExcel2])


# ## Creacion de consolidacion de CRM y Cierre

# In[33]:


# Creamos un arreglo con lo excel filtrados
carpeta = Path(rutaExcelCRM) 
archivo = list(carpeta.rglob("CRM*.xlsx")) + list(carpeta.glob("CRM*.xls")) # se filtra por CRM xlsx u xls
archivos_str = [str(a) for a in archivo] # retorna un array ['xxx','xxxx',...]de las ruta de los archivos
archivos_str


# In[34]:


# Usamos la funcion de ProcesarCRM
archivo_lectura = []
# archivo_cierre = []
for archivo  in archivos_str:
    df_1 = procesarCRM(archivo, hoja_principal="Base CRM", hoja_backup="CRM")
    if df_1 is None:
        continue
    archivo_lectura.append(df_1)

dfFinalConsolidadoCRM = pd.concat(archivo_lectura, ignore_index=True) # creamos un dataframe 


# In[35]:


# Procesamo para la creacion del consolidado Cierre
datosCierre = []
for archivo in archivos_str:
    # desarrollo de cierre
    claves = ['No.','Fecha','Asesor','Razon Social','RUC']
    pattern = "|".join(claves)
    dfCierre = pd.read_excel(archivo, sheet_name='CIERRE', header=None)

    # Buscar fila donde están los encabezados
    filaCierre = dfCierre.index[
        dfCierre.apply(lambda row: row.astype(str).str.contains(pattern, case=False).any(), axis=1)
    ][0]

    print("Fila encontrada:", filaCierre)

    # Cargar excel usando esa fila como encabezado
    df_final = pd.read_excel(
        archivo,
        sheet_name='CIERRE',
        header=filaCierre
    )

    datosCierre.append(df_final)

dfCierres = pd.concat(datosCierre, ignore_index=True)


# In[36]:


# Hacemos un cruce con el dataframe del consolidado de lista Precios
dfFinalCierre = pd.merge(dfCierres, dfFinalListaPrecios, left_on="Código", right_on="CÓDIGO", how='left')
dfFinalCierre = dfFinalCierre.drop(columns={'CÓDIGO','NOMBRE COMERCIAL','Costo','Costos','fecha'}).rename(columns={'COSTO COMPASS':'Costo'})
print(dfFinalCierre)


# ## Creamos los Excel

# In[37]:


carpeta = Path("Reportes")
carpeta.mkdir(exist_ok=True)


# In[38]:


# Creamos el excel de de lista precios.
ruta_Precios = carpeta / "ConsolidadoPrecios.xlsx"
with pd.ExcelWriter(ruta_Precios, engine="openpyxl") as writer:
    dfFinalListaPrecios.to_excel(writer, index=False)
    wb = writer.book
    ws = writer.sheets['Sheet1']

    for col in ['A',"B"]:
        for cell in ws[col]:
            if cell.value is not None:
                cell.number_format =  "@"

    for cell in ws["C"]:
        if cell.value is not None:
            cell.number_format = "#.##0"
    # Crear la tabla automáticamente
    max_row = ws.max_row
    max_col = ws.max_column
    ref = f"A1:{chr(64+max_col)}{max_row}"  # Rango dinámico según tus datos
    tabla = Table(displayName="TablaPrecios", ref=ref)

    # Estilo de tabla
    style = TableStyleInfo(name="TableStyleMedium9",
                           showFirstColumn=False,
                           showLastColumn=False,
                           showRowStripes=True,
                           showColumnStripes=True)
    tabla.tableStyleInfo = style
    ws.add_table(tabla)

     # Ajustar ancho de columnas al contenido
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener letra de la columna
        for cell in col:
            if cell.value:
                # Convertimos a str para calcular longitud
                max_length = max(max_length, len(str(cell.value)))
        # Un pequeño margen extra
        ws.column_dimensions[column].width = max_length + 2


# In[39]:


ruta_CierreCRM = carpeta / "ConsolidadoCierreCrm.xlsx"
with pd.ExcelWriter(ruta_CierreCRM, engine='openpyxl') as writer:

    dfFinalConsolidadoCRM.to_excel(writer, sheet_name="CRM", index=False)
    dfFinalCierre.to_excel(writer, sheet_name="CIERRE", index=False)

    wb = writer.book

    # =====================================================
    # HOJA 1
    # =====================================================
    ws1 = wb["CRM"]

    # Formatos
    for col in ['B', 'U']:
        for cell in ws1[col]:
            if cell.value:
                cell.number_format = "DD/MM/YYYY"

    for col in ['D', 'E', 'F']:
        for cell in ws1[col]:
            if cell.value:
                cell.number_format = "@"

    # Rango de tabla HOJA 1
    max_row = ws1.max_row
    max_col = ws1.max_column
    end_col = get_column_letter(max_col)
    ref = f"A1:{end_col}{max_row}"

    tabla1 = Table(displayName="TablaCRM_1", ref=ref)
    style1 = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    tabla1.tableStyleInfo = style1
    ws1.add_table(tabla1)

    # Autoajuste
    for col in ws1.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws1.column_dimensions[col[0].column_letter].width = max_length + 2

    # =====================================================
    # HOJA 2
    # =====================================================
    ws2 = wb["CIERRE"]

    max_row = ws2.max_row
    max_col = ws2.max_column
    end_col = get_column_letter(max_col)
    ref = f"A1:{end_col}{max_row}"

    tabla2 = Table(displayName="TablaCierre_1", ref=ref)
    style2 = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    tabla2.tableStyleInfo = style2
    ws2.add_table(tabla2)

    # Autoajuste
    for col in ws2.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws2.column_dimensions[col[0].column_letter].width = max_length + 2

